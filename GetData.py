#autor: victor fuentes toledo
#Version 1.1


import time
import os
from datetime import date
from openpyxl import load_workbook
import getpass
import warnings
from cryptography.utils import CryptographyDeprecationWarning
with warnings.catch_warnings():
    warnings.filterwarnings('ignore', category=CryptographyDeprecationWarning)
    import paramiko

ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())


host = '10.213.36.9'
name_excel='Metricas_API_Nozomi.xlsx'
ruta_de_files_nozomi='/home/administrador/Documentos/API_Nozomi/Output'

print("################################")
print("###   Fast Reader & Write   ####")
print("###   Data Querys   v1.1    ####")
print("################################\n")


username=input('introduce usuario: ')
password=getpass.getpass(prompt='Introduce credenciales: ')
#password=input('Introduce credenciales: ')

try:


    time.sleep(5)
    ssh.connect(host, username=username, password=password)
    sftp = ssh.open_sftp()
    os.system("cls")
    print('Connecting to ...'+ host)

    time.sleep(10)
    print('Connected to ...' + host+'\n')
    (stdin, stdout, stderr) = ssh.exec_command('ls '+ruta_de_files_nozomi)
    os.system("cls")
    cmd_output = stdout.read()
    data_utf8_ls=str(cmd_output.decode('utf-8'))
    print("Elige el fichero para recoger los datos: \n")

    print(data_utf8_ls)
    getdatafile = input("Introduce el nombre del fichero: ")

    os.system("cls")
    print("El fichero elegido -> ",getdatafile)
    (stdinn, stdoutt, stderrr) = ssh.exec_command('ls '+ruta_de_files_nozomi+'/'+str(getdatafile))
    cmd_output = stdoutt.read()
    data_utf8_checkfile=str(cmd_output.decode('utf-8'))

    if data_utf8_checkfile=='':
        print('El fichero introducido no existe..')
    else:
        directorio_local = os.getcwd()
        os.system("cls")
        print("Descargando datos y Escribiendo en Reporte..")

        'DESCARGA DEL FICHERO ELEGIDO'
        sftp.get(ruta_de_files_nozomi+'/'+getdatafile, os.path.join(directorio_local,getdatafile))

        today = date.today()
        ddmmyy = today.strftime("%d/%m/%Y")

        'CARGA DE EXCEL'
        try:
            wb = load_workbook(directorio_local+'\\'+name_excel)
            ws = wb.active
            'ELIGE LA HOJA PARA LEER'
            hoja_datos = wb['Datos']

            rellena_ultima_linea=hoja_datos.max_row+1

            f = open(getdatafile,"r")
            lines = f.readlines()

            for line in lines:

                hoja_datos.cell(row=rellena_ultima_linea, column=1,value=line.split(",")[1])
                hoja_datos.cell(row=rellena_ultima_linea, column=2,value=line.split(",")[2])
                hoja_datos.cell(row=rellena_ultima_linea, column=3,value = ddmmyy)
                rellena_ultima_linea=rellena_ultima_linea+1

            wb.save(directorio_local+'\\'+name_excel)
            os.system("cls")
            print("Fin del script")

        except:
            os.system("cls")
            print("Error de escritura -> Cierre el documento excel..",directorio_local+'\\'+name_excel)

except TimeoutError as timeouterror:
        print(timeouterror)
except paramiko.ssh_exception.AuthenticationException as ssherr:
        print(ssherr)
except paramiko.ssh_exception.NoValidConnectionsError as not_connection:
        print(not_connection)


os.system("pause")































































